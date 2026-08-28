from services.time_utils import utcnow
import csv
import hashlib
import hmac
import io
import json
import os
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

from flask import Blueprint, Response, abort, current_app, flash, jsonify, redirect, render_template, request, session, url_for
from sqlalchemy import text, func, or_
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

from db import db
from models.category.category import Category
from models.client.client import Client
from models.company.company import Company
from models.operations import BillingInvoice, OperationJob, SubscriptionEvent
from models.products.products import Product, ProductType
from models.supplier.supplier import Supplier
from models.user.user import User
from models.warehouse.warehouse import Warehouse
from models.warehouse_stock.warehouse_stock import WarehouseStock
from models.stock_movement.stock_movement import StockMovement
from models.retail import UnitOfMeasure, ProductUomConversion, ProductBarcode
from services.numeric import bounded_decimal, finite_int
from services.quantity import as_decimal, conversion_factor, non_negative_quantity, product_quantity
from services.csv_security import safe_csv_row
from services.validation import BusinessRuleError


operations_bp = Blueprint('operations_bp', __name__, url_prefix='/operations')


def _company():
    return Company.query.filter_by(id=session.get('company_id')).first_or_404()


def _parse_datetime(value):
    if not value:
        return None
    return datetime.fromisoformat(str(value).replace('Z', '+00:00')).replace(tzinfo=None)


@operations_bp.get('/health/live')
def health_live():
    return jsonify(status='ok', service='orbiserp', release=current_app.config.get('RELEASE_VERSION'))


@operations_bp.get('/health/ready')
def health_ready():
    checks = {'database': 'error', 'migrations': 'error', 'storage': 'error'}
    try:
        db.session.execute(text('SELECT 1'))
        checks['database'] = 'ok'
        revision = db.session.execute(text('SELECT version_num FROM alembic_version')).scalar_one()
        checks['migrations'] = (
            'ok' if revision == current_app.config['EXPECTED_SCHEMA_REVISION'] else 'outdated'
        )
    except Exception:
        current_app.logger.exception('Readiness check failed')
        db.session.rollback()
    storage_root = current_app.config.get('STORAGE_ROOT') or current_app.static_folder
    if storage_root and os.path.isdir(storage_root) and os.access(storage_root, os.R_OK | os.W_OK):
        checks['storage'] = 'ok'
    ready = all(value == 'ok' for value in checks.values())
    return jsonify(status='ready' if ready else 'unavailable', release=current_app.config.get('RELEASE_VERSION'), **checks), 200 if ready else 503


@operations_bp.route('/setup', methods=['GET', 'POST'])
def onboarding():
    company = _company()
    user = db.session.get(User, session.get('user_id'))
    if request.method == 'POST':
        company.name = (request.form.get('name') or company.name).strip()[:150]
        company.rnc = (request.form.get('rnc') or '').strip()[:20] or None
        company.email = (request.form.get('email') or '').strip()[:120] or None
        company.phone = (request.form.get('phone') or '').strip()[:20] or None
        company.address = (request.form.get('address') or '').strip() or None
        warehouse_name = (request.form.get('warehouse_name') or '').strip()
        if warehouse_name and not Warehouse.query.filter_by(company_id=company.id).first():
            db.session.add(Warehouse(name=warehouse_name[:150], location=company.address, is_main=True, company_id=company.id))
        company.onboarding_completed = True
        db.session.commit()
        flash('Configuración inicial completada. Tu empresa está lista para operar.', 'success')
        return redirect(url_for('dashboard_bp.dashboard'))
    return render_template('operations/setup.html', company=company, user=user)


@operations_bp.get('/billing')
def billing():
    company = _company()
    user = db.session.get(User, session.get('user_id'))
    invoices = BillingInvoice.query.filter_by(company_id=company.id).order_by(BillingInvoice.created_at.desc()).limit(24).all()
    return render_template(
        'operations/billing.html',
        company=company,
        invoices=invoices,
        user=user,
        billing_mode=current_app.config.get('BILLING_MODE', 'manual'),
    )


@operations_bp.post('/billing/webhook')
def billing_webhook():
    secret = current_app.config.get('BILLING_WEBHOOK_SECRET', '')
    if not secret:
        return jsonify(error='Webhook no configurado'), 503
    raw = request.get_data(cache=True)
    received = request.headers.get('X-Orbis-Signature', '')
    expected = hmac.new(secret.encode(), raw, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(received, expected):
        abort(401)
    payload = request.get_json(silent=True) or {}
    event_id = str(payload.get('id') or '')[:150]
    event_type = str(payload.get('type') or '')[:80]
    company_id = payload.get('company_id')
    if not event_id or not event_type or not isinstance(company_id, int):
        return jsonify(error='Evento incompleto'), 400
    if SubscriptionEvent.query.filter_by(event_id=event_id).first():
        return jsonify(status='duplicate'), 200
    event = SubscriptionEvent(event_id=event_id, event_type=event_type, company_id=company_id,
                              provider=str(payload.get('provider') or 'generic')[:40],
                              payload_hash=hashlib.sha256(raw).hexdigest())
    db.session.add(event)
    try:
        db.session.flush()
    except IntegrityError:
        db.session.rollback()
        return jsonify(status='duplicate'), 200
    company = db.session.get(Company, company_id)
    if not company:
        event.error = 'Empresa no encontrada'
        db.session.commit()
        return jsonify(error=event.error), 404
    data = payload.get('data') or {}
    try:
        if event_type in {'payment.succeeded', 'subscription.activated', 'subscription.renewed'}:
            plan = str(data.get('plan') or company.plan_name).upper()
            if plan not in {'BASIC', 'PRO', 'ULTRA'}:
                raise ValueError('Plan inválido')
            company.plan_name = plan
            company.plan_status = 'ACTIVE'
            company.status = True
            company.is_readonly = False
            company.billing_provider = event.provider
            company.billing_customer_id = str(data.get('customer_id') or '')[:120] or company.billing_customer_id
            company.billing_subscription_id = str(data.get('subscription_id') or '')[:120] or company.billing_subscription_id
            company.expiration_date = _parse_datetime(data.get('period_end')) or utcnow() + timedelta(days=30)
            company.grace_period_until = None
            external_id = str(data.get('invoice_id') or event_id)[:120]
            invoice = BillingInvoice.query.filter_by(external_id=external_id).first()
            if not invoice:
                invoice = BillingInvoice(company_id=company.id, external_id=external_id, provider=event.provider,
                                         status='PAID', plan_name=plan, amount=bounded_decimal(
                                             data.get('amount'), field_name='Monto de factura', places=2,
                                             minimum='0', maximum='9999999999.99', allow_blank=True, default='0',
                                         ),
                                         currency=str(data.get('currency') or 'DOP')[:3].upper(),
                                         period_start=_parse_datetime(data.get('period_start')),
                                         period_end=company.expiration_date, paid_at=utcnow())
                db.session.add(invoice)
        elif event_type == 'payment.failed':
            company.plan_status = 'PAST_DUE'
            company.grace_period_until = utcnow() + timedelta(days=3)
        elif event_type == 'subscription.cancelled':
            company.plan_status = 'CANCELLED'
            company.cancel_at_period_end = True
        else:
            raise ValueError('Tipo de evento no soportado')
        event.processed = True
        db.session.commit()
    except (ValueError, InvalidOperation) as exc:
        db.session.rollback()
        return jsonify(error=str(exc)), 400
    return jsonify(status='processed')


def _rows_from_bytes(filename, data):
    filename = (filename or '').lower()
    if filename.endswith('.csv'):
        try:
            content = data.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise ValueError('El CSV debe estar codificado en UTF-8.') from exc
        return list(csv.DictReader(io.StringIO(content)))
    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError('Falta openpyxl. Ejecuta pip install -r requirements.txt.') from exc
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or '').strip() for value in next(iterator)]
        except StopIteration:
            return []
        return [{headers[i]: (row[i] if i < len(row) and row[i] is not None else '') for i in range(len(headers))} for row in iterator]
    raise ValueError('Usa un archivo CSV o XLSX.')


def _uploaded_rows(upload):
    filename = (upload.filename or '').lower()
    data = upload.stream.read()
    if not filename.endswith(('.csv', '.xlsx')):
        raise ValueError('Usa un archivo CSV o XLSX.')
    return _rows_from_bytes(filename, data)


@operations_bp.route('/data', methods=['GET', 'POST'])
def data_center():
    company = _company()
    user = db.session.get(User, session.get('user_id'))
    report = []
    preview = []
    if request.method == 'POST':
        kind = request.form.get('kind')
        upload = request.files.get('file')
        if kind not in {'products', 'clients', 'suppliers', 'inventory'} or not upload:
            flash('Selecciona un tipo y un archivo CSV/XLSX válido.', 'danger')
            return redirect(request.url)
        try:
            rows = _uploaded_rows(upload)
        except ValueError as exc:
            flash(str(exc), 'danger')
            return redirect(request.url)
        if request.form.get('action') == 'preview':
            preview = rows[:20]
            return render_template('operations/data_center.html', company=company, user=user, report=report, preview=preview, preview_kind=kind)
        if len(rows) > 5000:
            flash('El archivo supera el máximo de 5,000 filas por importación.', 'danger')
            return redirect(request.url)
        job = OperationJob(
            company_id=company.id, user_id=user.id, job_type=f'IMPORT_{kind.upper()}',
            status='RUNNING', progress=5, total_rows=len(rows), started_at=utcnow(),
        )
        db.session.add(job)
        db.session.commit()
        job_id = job.id
        for number, row in enumerate(rows, 2):
            try:
                # A savepoint forces database constraints to be checked per row
                # without poisoning the complete import transaction. The import
                # remains atomic: any reported row rolls back every data change.
                with db.session.begin_nested():
                    if kind == 'products':
                        _import_product(company.id, row)
                    elif kind == 'clients':
                        _import_client(company.id, row)
                    elif kind == 'suppliers':
                        _import_supplier(company.id, row)
                    else:
                        _import_inventory(company.id, row)
            except IntegrityError:
                report.append(
                    f'Fila {number}: existe un valor duplicado o una relación inválida '
                    '(SKU, código de barras, correo, almacén o inventario).'
                )
            except (BusinessRuleError, ValueError, InvalidOperation) as exc:
                report.append(f'Fila {number}: {exc}')
            except SQLAlchemyError:
                current_app.logger.exception(
                    'Database import row failure company_id=%s kind=%s row=%s',
                    company.id, kind, number,
                )
                report.append(f'Fila {number}: la base de datos rechazó el registro.')
        if report:
            db.session.rollback()
            current_app.logger.warning(
                'Importación %s fallida company_id=%s filas=%s errores=%s detalle=%s',
                kind, company.id, len(rows), len(report), ' | '.join(report[:10])
            )
            job = db.session.get(OperationJob, job_id)
            job.status = 'FAILED'
            job.progress = 100
            job.processed_rows = max(len(rows) - len(report), 0)
            job.error_count = len(report)
            job.error_summary = '\n'.join(report[:50])[:5000]
            job.finished_at = utcnow()
            db.session.commit()
            flash(f'Importación detenida: corrige {len(report)} filas.', 'warning')
        else:
            current_app.logger.info(
                'Importación %s completada company_id=%s filas=%s', kind, company.id, len(rows)
            )
            job.status = 'COMPLETED'
            job.progress = 100
            job.processed_rows = len(rows)
            job.finished_at = utcnow()
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                job = db.session.get(OperationJob, job_id)
                job.status = 'FAILED'
                job.progress = 100
                job.processed_rows = 0
                job.error_count = 1
                job.error_summary = 'La base de datos rechazó la importación al confirmar los cambios.'
                job.finished_at = utcnow()
                db.session.commit()
                flash(
                    'No se importó ningún registro: la base de datos detectó un duplicado o una relación inválida.',
                    'danger',
                )
            else:
                flash(f'{len(rows)} registros importados correctamente.', 'success')
        return render_template('operations/data_center.html', company=company, user=user, report=report, preview=preview)
    return render_template('operations/data_center.html', company=company, user=user, report=report, preview=preview)


def _resolve_import_uom(company_id, value, field_name):
    raw = str(value or '').strip()
    if not raw:
        return None
    uom = UnitOfMeasure.query.filter(
        UnitOfMeasure.company_id == company_id,
        UnitOfMeasure.active.is_(True),
        or_(func.lower(UnitOfMeasure.name) == raw.lower(), func.lower(UnitOfMeasure.symbol) == raw.lower()),
    ).first()
    if not uom:
        raise ValueError(f'{field_name}: unidad de medida no encontrada ({raw})')
    return uom


def _import_uom_factor(raw_value, selected, base, field_name):
    if not selected or not base or selected.id == base.id:
        return Decimal('1')
    raw = str(raw_value or '').strip()
    if raw:
        try:
            return conversion_factor(raw, field_name)
        except ValueError as exc:
            raise ValueError(str(exc)) from exc

    selected_factor = conversion_factor(selected.factor_to_reference or 1, f'{field_name}: factor de unidad')
    base_factor = conversion_factor(base.factor_to_reference or 1, f'{field_name}: factor base')
    fallback = selected_factor / base_factor
    if fallback == 1:
        raise ValueError(f'{field_name}: indica el factor específico del producto (ej. Caja=24)')
    try:
        return conversion_factor(fallback, field_name)
    except ValueError as exc:
        raise ValueError(
            f'{field_name}: la relación entre unidades requiere más de 6 decimales; indica un factor explícito'
        ) from exc


def _import_product(company_id, row):
    name, sku = (row.get('name') or '').strip(), (row.get('sku') or '').strip()
    if not name or not sku:
        raise ValueError('name y sku son obligatorios')
    try:
        price = bounded_decimal(
            row.get('price'), field_name='price', places=2,
            minimum='0', maximum='99999999.99', allow_blank=True, default='0',
        )
        cost = bounded_decimal(
            row.get('cost'), field_name='cost', places=2,
            minimum='0', maximum='99999999.99', allow_blank=True, default='0',
        )
    except ValueError as exc:
        raise ValueError(str(exc)) from exc
    product = Product.query.filter_by(company_id=company_id, sku=sku).first()
    updating_existing = product is not None
    category = None
    category_name = str(row.get('category') or '').strip()
    if category_name:
        category = Category.query.filter(db.func.lower(Category.name) == category_name.lower(), Category.company_id == company_id).first()
        if not category:
            category = Category(name=category_name[:100], company_id=company_id, status=True)
            db.session.add(category)
            db.session.flush()
    type_name = str(row.get('type') or 'STOCKED').upper()
    if type_name not in ProductType.__members__:
        raise ValueError(f'tipo inválido: {type_name}')
    sale_mode = str(row.get('sale_mode') or 'UNIT').upper()
    tracking = str(row.get('tracking') or 'NONE').upper()
    if sale_mode not in {'UNIT', 'WEIGHT'}:
        raise ValueError('sale_mode debe ser UNIT o WEIGHT')
    if tracking not in {'NONE', 'LOT', 'SERIAL'}:
        raise ValueError('tracking debe ser NONE, LOT o SERIAL')
    base_uom = _resolve_import_uom(company_id, row.get('base_uom'), 'base_uom')
    sale_uom = _resolve_import_uom(company_id, row.get('sale_uom'), 'sale_uom') or base_uom
    purchase_uom = _resolve_import_uom(company_id, row.get('purchase_uom'), 'purchase_uom') or base_uom
    fractional_stock = (
        tracking != 'SERIAL'
        and (sale_mode == 'WEIGHT' or bool(base_uom and base_uom.allow_fraction))
    )
    try:
        warranty_days = finite_int(row.get('warranty_days') or '30', field_name='warranty_days')
        if not 1 <= warranty_days <= 3650:
            raise ValueError('warranty_days debe estar entre 1 y 3650')
        min_stock = non_negative_quantity(
            row.get('min_stock') or '0', 'min_stock', fractional=fractional_stock,
        )
        max_stock_raw = str(row.get('max_stock') or '').strip()
        max_stock = (
            non_negative_quantity(max_stock_raw, 'max_stock', fractional=fractional_stock)
            if max_stock_raw else None
        )
    except ValueError as exc:
        raise ValueError(str(exc)) from exc
    if max_stock is not None and max_stock < min_stock:
        raise ValueError('max_stock no puede ser menor que min_stock')
    if base_uom:
        for label, uom in (('sale_uom', sale_uom), ('purchase_uom', purchase_uom)):
            if uom and uom.category != base_uom.category:
                raise ValueError(f'{label} debe pertenecer a la misma categoría que base_uom')
    # Product photos are managed manually from the product form. Catalog
    # imports must never overwrite, detach or create product images.
    image_path = product.image_path if updating_existing else None
    image_url = product.image_url if updating_existing else None

    product_values = dict(
        name=name[:150], sku=sku[:50], description=str(row.get('description') or '').strip() or None,
        image_path=image_path, image_url=image_url,
        price=price, cost=cost, category_id=category.id if category else None,
        product_type=ProductType[type_name], company_id=company_id, status=True,
        brand=str(row.get('brand') or '').strip()[:100] or None,
        sale_mode=sale_mode, tracking=tracking, warranty_days=warranty_days,
        min_stock=min_stock, max_stock=max_stock,
        base_uom_id=base_uom.id if base_uom else None,
        sale_uom_id=sale_uom.id if sale_uom else None,
        purchase_uom_id=purchase_uom.id if purchase_uom else None,
    )
    if updating_existing:
        for field, value in product_values.items():
            setattr(product, field, value)
        # Reimportar un catálogo existente no debe duplicar inventario ni
        # conversiones. La foto manual se conserva sin cambios.
    else:
        product = Product(**product_values)
        db.session.add(product)
    db.session.flush()
    if base_uom and not updating_existing:
        conversion_map = {}
        for selected, field_name, allow_sale, allow_purchase in (
            (sale_uom, 'sale_uom_factor', True, False),
            (purchase_uom, 'purchase_uom_factor', False, True),
        ):
            if not selected or selected.id == base_uom.id:
                continue
            factor = _import_uom_factor(row.get(field_name), selected, base_uom, field_name)
            existing = conversion_map.get(selected.id)
            if existing and existing['factor'] != factor:
                raise ValueError(f'Los factores de venta/compra para {selected.name} no coinciden')
            data = existing or {'uom': selected, 'factor': factor, 'sale': False, 'purchase': False}
            data['sale'] = data['sale'] or allow_sale
            data['purchase'] = data['purchase'] or allow_purchase
            conversion_map[selected.id] = data
        for data in conversion_map.values():
            db.session.add(ProductUomConversion(
                company_id=company_id, product_id=product.id, uom_id=data['uom'].id,
                factor_to_base=data['factor'], allow_sale=data['sale'], allow_purchase=data['purchase'],
            ))
    barcode = str(row.get('barcode') or '').strip()
    if barcode:
        existing_barcode = ProductBarcode.query.filter_by(company_id=company_id, code=barcode).first()
        if existing_barcode:
            # Si products fue eliminado manualmente, pueden quedar barcodes huérfanos
            # apuntando a IDs que ya no existen. No deben bloquear una reimportación
            # válida del catálogo.
            owner = Product.query.filter_by(company_id=company_id, id=existing_barcode.product_id).first()
            if owner is None:
                db.session.delete(existing_barcode)
                db.session.flush()
                existing_barcode = None
            elif existing_barcode.product_id != product.id:
                raise ValueError(f'barcode duplicado en otro producto: {barcode}')
        if not existing_barcode:
            db.session.add(ProductBarcode(company_id=company_id, product_id=product.id, code=barcode[:120], barcode_type='INTERNAL', is_primary=True))
    try:
        stock = product_quantity(
            row.get('stock') or '0', 'stock', product=product, uom=base_uom, allow_zero=True,
        )
    except ValueError as exc:
        raise ValueError(str(exc)) from exc
    if stock and tracking in {'LOT', 'SERIAL'}:
        raise ValueError('para productos por lote/serie importa el producto con stock=0 y registra su trazabilidad después')
    if stock and product.product_type != ProductType.SERVICE and not updating_existing:
        warehouse = Warehouse.query.filter_by(company_id=company_id, is_main=True, status=True).first()
        if not warehouse:
            raise ValueError('crea un almacén principal antes de importar inventario')
        db.session.add(WarehouseStock(product_id=product.id, warehouse_id=warehouse.id,
                                      quantity=stock, company_id=company_id))
        db.session.add(StockMovement(
            company_id=company_id, user_id=session.get('user_id'), product_id=product.id, warehouse_id=warehouse.id,
            movement_type='IN', quantity=stock, reason='Inventario inicial por importación'
        ))


def _import_client(company_id, row):
    name = (row.get('name') or '').strip()
    if not name:
        raise ValueError('name es obligatorio')
    db.session.add(Client(name=name[:150], email=(row.get('email') or '').strip()[:120] or None,
                          phone=(row.get('phone') or '').strip()[:20] or None,
                          status=(row.get('status') or 'Lead')[:50], company_id=company_id))


def _import_supplier(company_id, row):
    name = (row.get('name') or '').strip()
    if not name:
        raise ValueError('name es obligatorio')
    db.session.add(Supplier(name=name[:150], email=(row.get('email') or '').strip()[:120] or None,
                            phone=(row.get('phone') or '').strip()[:50] or None, company_id=company_id))


def _import_inventory(company_id, row):
    sku = str(row.get('sku') or '').strip()
    warehouse_name = str(row.get('warehouse') or '').strip()
    mode = str(row.get('mode') or 'SET').strip().upper()
    if not sku or not warehouse_name:
        raise ValueError('sku y warehouse son obligatorios')
    if mode not in {'SET', 'ADD'}:
        raise ValueError('mode debe ser SET o ADD')
    product = Product.query.filter_by(company_id=company_id, sku=sku, status=True).filter(Product.archived_at.is_(None)).first()
    if not product or product.product_type == ProductType.SERVICE:
        raise ValueError(f'producto de inventario no encontrado: {sku}')
    try:
        qty = product_quantity(
            row.get('quantity') or '0', 'quantity', product=product, uom=product.base_uom, allow_zero=True,
        )
    except ValueError as exc:
        raise ValueError(str(exc)) from exc
    warehouse = Warehouse.query.filter(
        Warehouse.company_id == company_id, Warehouse.status.is_(True), func.lower(Warehouse.name) == warehouse_name.lower()
    ).first()
    if not warehouse:
        raise ValueError(f'almacén no encontrado: {warehouse_name}')
    stock = WarehouseStock.query.filter_by(company_id=company_id, product_id=product.id, warehouse_id=warehouse.id).with_for_update().first()
    before = as_decimal(stock.quantity) if stock else Decimal('0')
    after = before + qty if mode == 'ADD' else qty
    if not stock:
        stock = WarehouseStock(company_id=company_id, product_id=product.id, warehouse_id=warehouse.id, quantity=after)
        db.session.add(stock)
    else:
        stock.quantity = after
    delta = after - before
    if delta:
        db.session.add(StockMovement(company_id=company_id, user_id=session.get('user_id'), product_id=product.id, warehouse_id=warehouse.id,
                                     movement_type='IN' if delta > 0 else 'OUT', quantity=abs(delta),
                                     reason=f'Importación inventario ({mode})'))


@operations_bp.get('/data/export/<kind>.csv')
def export_data(kind):
    company_id = session.get('company_id')
    output = io.StringIO()
    writer = csv.writer(output)
    if kind == 'products':
        writer.writerow(['name', 'sku', 'description', 'price', 'cost', 'type', 'category', 'brand', 'sale_mode', 'tracking', 'warranty_days', 'min_stock', 'max_stock', 'base_uom', 'sale_uom', 'sale_uom_factor', 'purchase_uom', 'purchase_uom_factor', 'barcode', 'stock'])
        for p in Product.query.filter_by(company_id=company_id).order_by(Product.name).all():
            sale_factor = next((c.factor_to_base for c in p.uom_conversions if p.sale_uom_id and c.uom_id == p.sale_uom_id), 1 if p.sale_uom_id == p.base_uom_id else '')
            purchase_factor = next((c.factor_to_base for c in p.uom_conversions if p.purchase_uom_id and c.uom_id == p.purchase_uom_id), 1 if p.purchase_uom_id == p.base_uom_id else '')
            writer.writerow(safe_csv_row([p.name, p.sku, p.description or '', p.price, p.cost, p.product_type.name, p.category.name if p.category else '', p.brand or '', p.sale_mode, p.tracking, p.warranty_days, p.min_stock, p.max_stock if p.max_stock is not None else '', p.base_uom.symbol if p.base_uom else '', p.sale_uom.symbol if p.sale_uom else '', sale_factor, p.purchase_uom.symbol if p.purchase_uom else '', purchase_factor, next((b.code for b in p.barcodes if b.is_primary), ''), p.total_stock]))
    elif kind == 'clients':
        writer.writerow(['name', 'email', 'phone', 'status'])
        for c in Client.query.filter_by(company_id=company_id).order_by(Client.name).all():
            writer.writerow(safe_csv_row([c.name, c.email or '', c.phone or '', c.status]))
    elif kind == 'suppliers':
        writer.writerow(['name', 'email', 'phone'])
        for s in Supplier.query.filter_by(company_id=company_id).order_by(Supplier.name).all():
            writer.writerow(safe_csv_row([s.name, s.email or '', s.phone or '']))
    else:
        abort(404)
    return Response('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename=orbiserp_{kind}.csv'})


@operations_bp.get('/data/template/<kind>.csv')
def download_import_template(kind):
    """Plantillas mínimas para evitar encabezados inválidos al importar."""
    headers = {
        'products': ['name', 'sku', 'description', 'price', 'cost', 'type', 'category', 'brand', 'sale_mode', 'tracking', 'warranty_days', 'min_stock', 'max_stock', 'base_uom', 'sale_uom', 'sale_uom_factor', 'purchase_uom', 'purchase_uom_factor', 'barcode', 'stock'],
        'clients': ['name', 'email', 'phone', 'status'],
        'suppliers': ['name', 'email', 'phone'],
        'inventory': ['sku', 'warehouse', 'quantity', 'mode'],
    }
    if kind not in headers:
        abort(404)
    output = io.StringIO()
    csv.writer(output).writerow(headers[kind])
    return Response('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename=plantilla_{kind}.csv'})
